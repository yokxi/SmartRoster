from flask import Flask, render_template, request, send_file
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import io
import calendar
import locale
from collections import defaultdict

try:
    locale.setlocale(locale.LC_TIME, 'en_US.UTF-8')
except locale.Error:
    print("Locale 'en_US.UTF-8' not found, using default 'C'.")
    locale.setlocale(locale.LC_TIME, 'C')

print("Starting SmartRoster Server (Daily Roster v)...")

app = Flask(__name__)

def calculate_shift_duration(time_slot_str):
    try:
        t1_str, t2_str = time_slot_str.strip().split(' - ')
        t1 = datetime.strptime(t1_str, '%H:%M')
        t2 = datetime.strptime(t2_str, '%H:%M')
        duration = (t2 - t1).total_seconds() / 3600
        return duration
    except Exception:
        return 0

def find_suitable_employee(employees, day_name, already_in_shift, current_week):
    best_candidate = None
    best_score = float('inf')

    for p in employees:
        if p["name"] in already_in_shift:
            continue
        
        current_score = p["hours_worked"]
        
        if day_name in ["Saturday", "Sunday"] and p["last_weekend_worked"] == current_week - 1:
            current_score += 1000
        if p["hours_worked"] >= p["target_monthly_hours"]:
            current_score += 500
        
        if current_score < best_score:
            best_score = current_score
            best_candidate = p
            
    return best_candidate

def generate_and_transform_schedules(employees, schedules, year, month):
    print(f"Starting schedule generation for {month}/{year}...")
    
    month_cal = calendar.monthcalendar(year, month)
    day_names = list(calendar.day_name)
    day_abbrs = [day.capitalize()[:3] for day in calendar.day_abbr]
    
    store_view = {}
    employee_view = defaultdict(list)
    
    for p in employees:
        employee_view[p["name"]] = []
        
    for week_index, week in enumerate(month_cal):
        for day_index, day_date in enumerate(week):
            
            if day_date == 0:
                continue
                
            current_day_name = day_names[day_index].capitalize()
            current_day_abbr = day_abbrs[day_index]
            
            if current_day_name in schedules:
                for shift_info in schedules[current_day_name]:
                    current_shift = shift_info["shift"]
                    required_staff = shift_info["required_staff"]
                    shift_duration = calculate_shift_duration(current_shift)
                    
                    assigned_names_list = []
                    
                    for _ in range(required_staff):
                        chosen_employee = find_suitable_employee(
                            employees, current_day_name, assigned_names_list, week_index
                        )
                        
                        if chosen_employee:
                            chosen_name = chosen_employee["name"]
                            assigned_names_list.append(chosen_name)
                            
                            employee_view[chosen_name].append({
                                "day": day_date,
                                "day_name": current_day_abbr,
                                "shift": current_shift,
                                "duration": shift_duration
                            })
                            
                            for p in employees:
                                if p["name"] == chosen_name:
                                    p["hours_worked"] += shift_duration
                                    if current_day_name in ["Saturday", "Sunday"]:
                                        p["last_weekend_worked"] = week_index
                                    break
                        else:
                            assigned_names_list.append("???")
                    
                    store_view[(day_date, current_shift)] = ", ".join(assigned_names_list)

    print("...Schedules calculated and transformed.")
    print("--- Monthly Hours Summary ---")
    for p in employees:
        print(f"{p['name']}: {p['hours_worked']} hours / {p['target_monthly_hours']} target")
    
    for name in employee_view:
        employee_view[name].sort(key=lambda shift: shift["day"])
        
    return store_view, employee_view

def create_daily_roster_excel(employee_view, employees, year, month):
    print("Creating 'Daily Roster' Excel file with colored headers...")
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    month_name = calendar.month_name[month].capitalize()
    ws.title = f"Roster {month_name} {year}"
    
    base_header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    weekend_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
    off_font = Font(color="888888")
    total_font = Font(bold=True)
    
    employee_header_colors = [
        "004D40", 
        "4E342E", 
        "1A237E", 
        "BF360C", 
        "0D47A1", 
        "33691E", 
    ]

    ws.cell(row=1, column=1, value="Date").fill = base_header_fill
    ws.cell(row=1, column=1, value="Date").font = header_font
    ws.cell(row=1, column=2, value="Day").fill = base_header_fill
    ws.cell(row=1, column=2, value="Day").font = header_font
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10

    name_to_col_map = {}
    current_col = 3
    color_index = 0
    
    for p in employees:
        name = p["name"]
        
        color_hex = employee_header_colors[color_index % len(employee_header_colors)]
        employee_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

        header_cell = ws.cell(row=1, column=current_col, value=name)
        header_cell.fill = employee_fill
        header_cell.font = header_font
        
        ws.column_dimensions[openpyxl.utils.get_column_letter(current_col)].width = 18
        name_to_col_map[name] = current_col
        current_col += 1
        color_index += 1

    days_in_month = calendar.monthrange(year, month)[1]
    day_abbrs = [day.capitalize()[:3] for day in calendar.day_abbr]
    
    current_row = 2
    
    for day_date in range(1, days_in_month + 1):
        
        day_of_week_index = datetime(year, month, day_date).weekday()
        day_name_abbr = day_abbrs[day_of_week_index]
        
        ws.cell(row=current_row, column=1, value=f"{day_date} {month_name[:3]}")
        ws.cell(row=current_row, column=2, value=day_name_abbr)
        
        is_weekend = (day_of_week_index >= 5)
        
        for col_idx in range(1, 3): 
             if is_weekend:
                ws.cell(row=current_row, column=col_idx).fill = weekend_fill

        for employee_name, col_idx in name_to_col_map.items():
            
            shift_for_day = None
            for shift in employee_view[employee_name]:
                if shift["day"] == day_date:
                    shift_for_day = shift["shift"]
                    break
            
            cell = ws.cell(row=current_row, column=col_idx)
            if shift_for_day:
                cell.value = shift_for_day
            else:
                cell.value = "OFF"
                cell.font = off_font
            
            if is_weekend:
                 cell.fill = weekend_fill
                
        current_row += 1
        
    current_row += 1
    ws.cell(row=current_row, column=2, value="Total Hours:").font = total_font
    
    for p in employees:
        col_idx = name_to_col_map[p["name"]]
        ws.cell(row=current_row, column=col_idx, value=p["hours_worked"]).font = total_font

    print("...'Daily Roster' Excel file created in memory.")

    memory_file = io.BytesIO()
    wb.save(memory_file)
    memory_file.seek(0)
    
    return memory_file, f"roster_{month_name.lower()}_{year}.xlsx"

def parse_employees(text_input):
    employees = []
    try:
        for line in text_input.strip().split('\n'):
            line = line.strip()
            if not line: continue
            parts = line.split(',')
            name = parts[0].strip()
            monthly_hours = int(parts[1].strip())
            
            employees.append({
                "name": name,
                "target_monthly_hours": monthly_hours,
                "hours_worked": 0,
                "last_weekend_worked": -1
            })
    except Exception as e:
        print(f"Error parsing employees: {e}")
        return []
    return employees

def parse_schedules(text_input):
    schedules = {}
    try:
        for line in text_input.strip().split('\n'):
            line = line.strip()
            if not line: continue
            parts = line.split(',')
            day = parts[0].strip()
            shift = parts[1].strip()
            staff = int(parts[2].strip())
            
            if day not in schedules:
                schedules[day] = []
            schedules[day].append({
                "shift": shift,
                "required_staff": staff
            })
    except Exception as e:
        print(f"Error parsing schedules: {e}")
        return {}
    return schedules

@app.route("/")
def index():
    return render_template('index.html')

@app.route("/generate", methods=['POST'])
def handle_generation():
    print("Received data from form!")
    
    employee_text = request.form['employees_input']
    schedule_text = request.form['schedule_input']
    month = int(request.form['month'])
    year = int(request.form['year'])
    
    employee_list = parse_employees(employee_text)
    schedule_coverage = parse_schedules(schedule_text)
    
    store_view, employee_view = generate_and_transform_schedules(
        employee_list, schedule_coverage, year, month
    )
    
    file_to_download, file_name = create_daily_roster_excel(
        employee_view, employee_list, year, month
    )
    
    return send_file(
        file_to_download,
        download_name=file_name,
        as_attachment=True,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

if __name__ == "__main__":
    app.run(debug=True)